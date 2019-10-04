# Reading an excel file using Python
import xlrd
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
# from openpyxl import Workbook

# Give the location of the file
takeoff_doc = './/Document #2 (After).xlsx'
estimate_doc = './COST SPREADSHEET.xlsx'
result_doc='./COST SPREADSHEET - Modified.xlsx'
wb=load_workbook(estimate_doc)

# To open Workbook
takeoff = xlrd.open_workbook(takeoff_doc)
estimate = xlrd.open_workbook(estimate_doc)

source_sheet = takeoff.sheet_by_index(0)
target_sheet = estimate.sheet_by_name('ESTIMATE')
col = 0


class DataBlock:
    def __init__(self, name, start):
        self.sourceName = name
        self.targetName = self.whatAmI(name)
        self.start = start
        self.end = -1
        self.target = -1

    def __str__(self):
        return self.sourceName+": "+str(self.start+1)+" -> "+str(self.end+1)+"  ("+str(self.target+1)+")"
    
    def whatAmI(self, name):
        if name=="elevator pit":
            return "elevator pit \" walls"
        elif name=="slab":
            return "slab step"
        else:
            return name

# SEARCH FOR EACH FLOOR
floors = []
for row in range(source_sheet.nrows):
    col = 0
    # print(source_sheet.cell(row,col).value)
    if source_sheet.cell(row, col).value.lower() == 'foundation':
        floors.append(row)
    elif 'floor' in str(source_sheet.cell(row, col).value).lower() and str(source_sheet.cell(row-1, col).value) == "":
        floors.append(row)


print("SOURCE FLOORS:")
for f in range(len(floors)):
     print("  FLOOT "+str(f+1)+": "+str(floors[f]+1))

# GET TARGET SHEET FLOORS
target_locations={}
for row in range(target_sheet.nrows):
    val = str(target_sheet.cell(row, col).value).lower().strip()
    if (len(val)<10 and val[-5:]=="level") or val=="foundation" or val=="main roof":
        target_locations.update({len(target_locations): row})


print("TARGET FLOORS:")
for t in target_locations.keys():
    print("  FLOOR "+str(t+1)+": "+str(target_locations[t]+1))

# ITERATE THROUGH FLOORS/SECTIONS
inserts=0
for floor in range(len(floors)):
    col=0
    categories = []
    reinforcing = []
    print("---- BEGIN TRANSFER OF FLOOR "+str(floor+1)+" ----")

    # SEARCH FOR START OF EACH CATEGORY
    if floor<len(floors)-1:
        floorEnd=floors[floor+1]
    else:
        floorEnd= source_sheet.nrows
    print("  - SEARCHING SOURCE AT "+str(floors[floor]+1)+" - "+str(floorEnd)+" FOR CATEGORIES")

    rangeVals = range(floors[floor]+1,floorEnd)
    for row in rangeVals:
        val = str(source_sheet.cell(row, col).value)
        if (val.isupper() and not any(char.isdigit() for char in val)):
            if val.lower().strip() == 'reinforcing':
                reinforcing.append(row)
            else:
                categories.append(DataBlock(val.lower(), row+2))


    # SEARCH FOR END OF EACH CATEGORY
    for index in range(len(categories)):
        # SET SIZE OF LOOP
        if len(categories) < index:
            rangeVals = range(categories[index].start, categories[index+1].start-1)
        else:
            rangeVals = range(categories[index].start, floorEnd)

        for row in rangeVals:
            val = str(source_sheet.cell(row, col).value)
            if len(val) == 0:
                categories[index].end = row-1
                break 

    # FIND CATEGORY LOCATION ON TARGET SHEET
    for index in range(len(categories)):
        # FIND START POINT ON ESTIMATE
        for row in range(target_locations[floor], target_locations[floor+1]):
            val = str(target_sheet.cell(row, col).value).lower().strip()
            # print(str(row)+") "+val+"-"+categories[index].name+"  "+str(val == categories[index].name))
            if val == categories[index].targetName:
                categories[index].target=row+2
                break

    print("CATEGORIES (SOURCE):")
    for cat in categories:
        print("  - "+str(cat))

    # GET ESTIMATE FILE WITH OPENPYXL
    estimate_sheets=wb.sheetnames
    if "ESTIMATE" not in estimate_sheets:
        print ("Sheet with name 'ESTIMATE' is missing from Estimate workbook")
        exit(1)

    print("INSERTING VALUES:")
    worksheet=wb['ESTIMATE']
    for cat in categories:
        if cat.target <= 0:
            print("  - "+str(cat.sourceName)+" -> NOT FOUND ON TARGET")
        else:
            print("  - "+str(cat.sourceName))
            for source_row in range(cat.start, cat.end):
                # FILL IN THE BLANK ROW BEFORE ADDING NEW ONES
                if source_row>cat.start:
                    worksheet.insert_rows(cat.target+inserts)

                print("    - "+str(cat.target+inserts)+"] "+str(source_sheet.cell(source_row,0).value))                    
                # COPY EACH COLUMN OVER
                source_col_mapping={0:0, 1:1, 2:2, 3:4, 4:5} #{target:source}
                for target_col in range(0,9):
                    
                    # SET BORDER
                    thin = Side(border_style="thin", color="000000")
                    worksheet.cell(column=target_col+1, row=cat.target+inserts).border=Border(top=thin, left=thin, bottom=thin, right=thin)
                    
                    #SET ALIGNMENT
                    if target_col==0 or target_col==4:
                        align="left"
                    elif target_col==3:
                        align="right"
                    else:
                        align="center"
                    worksheet.cell(column=target_col+1, row=cat.target+inserts).alignment=Alignment(horizontal=align)

                    #SET COLORS
                    if target_col==3:
                        worksheet.cell(column=target_col+1, row=cat.target+inserts).fill=PatternFill("solid",fgColor="E6B8B7")
                        
                    #SET COPIED & CALCULATED VALUES
                    if target_col in source_col_mapping:
                        source_col=source_col_mapping[target_col]
                        val=source_sheet.cell(source_row,source_col).value
                        worksheet.cell(column=target_col+1, row=cat.target+inserts, value=val)
                    elif target_col==5:
                        #TODO get the waste value
                        worksheet.cell(column=target_col+1, row=cat.target+inserts, value=0)
                    elif target_col==6:
                        val="T1"
                        worksheet.cell(column=target_col+1, row=cat.target+inserts, value=val)
                    elif target_col==8:
                        val="T2"
                        worksheet.cell(column=target_col+1, row=cat.target+inserts, value=val)

                # DONT ADD ANOTHER ROW ON THE LAST ELEMENT OF THE CATEGORY
                if source_row<cat.end-1:
                    inserts+=1

wb.save(result_doc)